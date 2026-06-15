import openpyxl
import os
import re

folder = r'c:\Users\User\Downloads\新增資料夾'

def test_file(filename):
    fpath = os.path.join(folder, filename)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    sheets = wb.sheetnames
    
    # Find monthly schedule sheet
    schedule_sheet = next((s for s in sheets if re.search(r'\d{6}', s)), None)
    if not schedule_sheet:
        print(f'{filename} -> No schedule sheet found!')
        return
        
    ws = wb[schedule_sheet]
    print(f'\n--- Testing {filename} (Sheet: {schedule_sheet}) ---')
    
    # Load all cells
    data = []
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        data.append(row_vals)
        
    # Find start col for dates (Row 2, index 1)
    date_start_col = -1
    dates = []
    if len(data) > 1:
        for c in range(len(data[1])):
            val = data[1][c]
            # Verify if it's a real date (date serial code > 40000 for dates after 2010, or is datetime)
            is_date = False
            if hasattr(val, 'month'):
                is_date = True
            elif isinstance(val, (int, float)) and val > 40000:
                is_date = True
                
            if is_date:
                date_start_col = c
                break
                
        if date_start_col != -1:
            for c in range(date_start_col, len(data[1])):
                val = data[1][c]
                if val:
                    dates.append((c, val))
                    
    print(f'Date start column: {date_start_col} (Col name: {openpyxl.utils.get_column_letter(date_start_col+1)})')
    print(f'Found {len(dates)} date columns.')
    
    # Parse employees
    employees = []
    if date_start_col != -1:
        for r in range(2, len(data)):
            row = data[r]
            name = row[0]
            shift = str(row[1] or '').strip()
            
            # Check name and shift pattern
            if name and re.search(r'\d{2}:\d{2}', shift):
                work_days = float(row[date_start_col - 5] or 0)
                off_count = float(row[date_start_col - 4] or 0)
                pto_count = float(row[date_start_col - 3] or 0)
                pto_al_count = float(row[date_start_col - 2] or 0)
                loa_count = float(row[date_start_col - 1] or 0)
                
                employees.append({
                    'name': name,
                    'shift': shift,
                    'work_days': work_days,
                    'off': off_count,
                    'pto': pto_count,
                    'pto_al': pto_al_count,
                    'loa': loa_count
                })
                
    print(f'Parsed {len(employees)} employees:')
    for emp in employees:
        print(f"  {emp['name']}: shift={emp['shift']}, work_days={emp['work_days']}, off={emp['off']}, pto={emp['pto']}, pto_al={emp['pto_al']}, loa={emp['loa']}")
        
test_file('2026_排班表_May.xlsx')
test_file('202605.xlsx')
