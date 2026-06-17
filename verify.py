import os
import sys
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

folder = r'c:\Users\User\Downloads\新增資料夾'

# Canonical names
CANONICAL_NAMES = [
    'Alex Chen', 'Amber Wang', 'Evan Liu', 'Howard Chen', 
    'Jacky Lee', 'Jian Kai Ding', 'Molly Song', 'Rex Liao', 'Sherry Lin'
]

def get_canonical_name(name):
    if not name:
        return ''
    name_clean = str(name).strip().replace(' ', '').lower()
    if ('ding' in name_clean or 'din' in name_clean) and ('kai' in name_clean or 'jian' in name_clean or 'jlian' in name_clean):
        return 'Jian Kai Ding'
    
    # Try match
    for cname in CANONICAL_NAMES:
        c_clean = cname.replace(' ', '').lower()
        if name_clean == c_clean:
            return cname
        
        # Check parts
        c_parts = cname.lower().split()
        n_parts = str(name).lower().strip().split()
        if c_parts[-1] == n_parts[-1]:  # Same last name
            # Check overlap
            c_first = c_parts[0]
            n_first = n_parts[0]
            if c_first in n_first or n_first in c_first:
                return cname
    return str(name).strip()

def to_float(val):
    if not val:
        return 0.0
    val_str = str(val).strip()
    import re
    m = re.match(r'^[-+]?[0-9]*\.?[0-9]+', val_str)
    if m:
        return float(m.group(0))
    return 0.0

def test_verification():
    # 1. Load data
    leave_wb = openpyxl.load_workbook(os.path.join(folder, '2026_請假申請表_Soluto_&_Care_May.xlsx'), data_only=True)
    ot_wb = openpyxl.load_workbook(os.path.join(folder, '2026_加班申請表_Soluto_&_Care_May.xlsx'), data_only=True)
    sched_wb = openpyxl.load_workbook(os.path.join(folder, '2026_排班表_May.xlsx'), data_only=True)
    target_wb = openpyxl.load_workbook(os.path.join(folder, '2026.5總請假統計.xlsx'), data_only=True)

    # Parse schedule employees
    sched_sheet = sched_wb['202605']
    sched_employees = []
    for r in range(3, sched_sheet.max_row+1):
        val = sched_sheet.cell(r, 1).value
        shift = sched_sheet.cell(r, 2).value
        if val and shift and any(c.isdigit() for c in str(shift)):
            sched_employees.append(get_canonical_name(val))
    sched_employees = sorted(list(set(sched_employees)))
    print('Schedule employees:', sched_employees)

    # Parse combined data from 班表相關申請表單(1-737).xlsx
    comb_wb = openpyxl.load_workbook(os.path.join(folder, '班表相關申請表單(1-737).xlsx'), data_only=True)
    comb_sheet = comb_wb['Sheet1']

    def normalize_leave_type(ltype):
        if not ltype:
            return ''
        ltype = str(ltype).strip()
        if '特別休假' in ltype or 'Annual Leave' in ltype or ('PTO' in ltype and 'PTO-AL' not in ltype):
            return 'Annual Leave特別休假'
        if '亞勝' in ltype or 'Asurion Leave' in ltype or 'PTO-AL' in ltype:
            return 'Asurion Leave亞勝假期'
        if '生理假' in ltype or 'Menstrual' in ltype:
            return 'Menstrual Leave病假（生理假)'
        if '病假' in ltype or 'Sick' in ltype:
            return 'Sick Leave病假'
        if '事假' in ltype or 'Personal' in ltype:
            return 'Personal Leave事假'
        if '金融市場' in ltype:
            return 'Official Leave公假（金融市場常識與職業道德考試）'
        if '財產保險' in ltype:
            return 'Official Leave公假（財產保險業務員資格證照考試)'
        if '健檢' in ltype:
            return 'Official Leave公假（健檢)'
        if '公假' in ltype or 'Official' in ltype:
            return 'Official Leave公假'
        if '婚假' in ltype or 'Marriage' in ltype:
            return 'Marriage Leave婚假'
        if '喪假' in ltype or 'Bereavement' in ltype:
            return 'Bereavement Leave喪假'
        if '家庭' in ltype or 'Family' in ltype:
            return 'Family Care Leave家庭照顧假'
        return ltype

    leave_records = []
    ot_records = []

    for r in range(2, comb_sheet.max_row+1):
        applicant = comb_sheet.cell(r, 6).value
        form_type = comb_sheet.cell(r, 7).value
        if not applicant or not form_type:
            continue
            
        name = get_canonical_name(applicant)
        form_type = str(form_type).strip()
        
        # Overtime
        if form_type.startswith('加班') or form_type == '加班申請表':
            date_val = comb_sheet.cell(r, 15).value # Column O
            hours = comb_sheet.cell(r, 19).value # Column S
            if date_val and hours:
                if isinstance(date_val, datetime):
                    dt_str = date_val.strftime('%Y-%m-%d')
                    m = date_val.month
                    y = date_val.year
                else:
                    dt_str = str(date_val).split()[0]
                    pts = dt_str.split('-')
                    y = int(pts[0])
                    m = int(pts[1])
                
                if y == 2026 and m == 5:
                    ot_records.append({'name': name, 'date': dt_str, 'hours': to_float(hours)})
                    
        # Leave
        elif (form_type.startswith('請假') or form_type.startswith('長假') or form_type.startswith('公假')) and '(事前申請)' not in form_type:
            ltype = ''
            start_date = None
            days = 0
            
            if form_type.startswith('請假'):
                ltype = normalize_leave_type(comb_sheet.cell(r, 8).value) # Column H
                start_date = comb_sheet.cell(r, 9).value # Column I
                days = comb_sheet.cell(r, 11).value # Column K
            elif form_type.startswith('長假'):
                ltype = normalize_leave_type(comb_sheet.cell(r, 29).value) # Column AC
                start_date = comb_sheet.cell(r, 26).value # Column Z
                days = comb_sheet.cell(r, 28).value # Column AB
            elif form_type.startswith('公假'):
                ltype = 'Official Leave公假'
                start_date = comb_sheet.cell(r, 30).value # Column AD
                days = 1
                
            if ltype and start_date:
                if not isinstance(start_date, datetime):
                    try:
                        start_date = datetime.strptime(str(start_date).split()[0], '%Y-%m-%d')
                    except Exception:
                        start_date = None
                
                if start_date:
                    from datetime import timedelta
                    remaining_days = to_float(days)
                    curr_date = start_date
                    while remaining_days > 0:
                        day_val = min(remaining_days, 1.0)
                        y = curr_date.year
                        m = curr_date.month
                        if y == 2026 and m <= 5:
                            leave_records.append({'name': name, 'type': ltype, 'month': m, 'days': day_val})
                        remaining_days -= day_val
                        curr_date += timedelta(days=1)

    # Calculate leave stats for May
    may_leave_stats = {}
    for r in leave_records:
        if r['month'] == 5:
            may_leave_stats.setdefault(r['name'], {})
            may_leave_stats[r['name']][r['type']] = may_leave_stats[r['name']].get(r['type'], 0.0) + r['days']

    print('\nCalculated May Leave Stats:')
    for name in sorted(may_leave_stats.keys()):
        print(f'  {name}: {may_leave_stats[name]}')

    # Compare with 2026.5總請假統計.xlsx May sheet
    print('\nComparing with Target May sheet:')
    target_may_sheet = target_wb['May']
    for r in range(4, target_may_sheet.max_row+1):
        emp_name = target_may_sheet.cell(r, 3).value
        if emp_name:
            cname = get_canonical_name(emp_name)
            target_vals = {
                'Annual Leave特別休假': float(target_may_sheet.cell(r, 4).value or 0),
                'Asurion Leave亞勝假期': float(target_may_sheet.cell(r, 5).value or 0),
                'Sick Leave病假': float(target_may_sheet.cell(r, 6).value or 0),
                'Menstrual Leave病假（生理假)': float(target_may_sheet.cell(r, 7).value or 0),
                'Personal Leave事假': float(target_may_sheet.cell(r, 8).value or 0),
                'Official Leave公假': float(target_may_sheet.cell(r, 9).value or 0),
                'Marriage Leave婚假': float(target_may_sheet.cell(r, 10).value or 0),
                'Bereavement Leave喪假': float(target_may_sheet.cell(r, 11).value or 0),
                'Family Care Leave家庭照顧假': float(target_may_sheet.cell(r, 12).value or 0),
            }
            calc_vals = may_leave_stats.get(cname, {})
            diffs = []
            for k, tv in target_vals.items():
                cv = calc_vals.get(k, 0.0)
                if k == 'Official Leave公假':
                    cv = sum(v for kt, v in calc_vals.items() if 'Official Leave公假' in kt)
                if cv != tv:
                    diffs.append(f'{k}: calc={cv} vs target={tv}')
            if diffs:
                print(f'  [DIFF] {cname}: {", ".join(diffs)}')
            else:
                print(f'  [OK] {cname} matches target May sheet!')

    # Calculate OT stats
    calculated_ot = {}
    for r in ot_records:
        calculated_ot.setdefault(r['name'], {})
        calculated_ot[r['name']][r['date']] = calculated_ot[r['name']].get(r['date'], 0.0) + r['hours']

    print('\nCalculated OT Stats:')
    for name in sorted(calculated_ot.keys()):
        print(f'  {name}: {calculated_ot[name]}')

    # Compare OT with 2026_加班申請表_Soluto_&_Care_May.xlsx sheet 加班整理
    print('\nComparing OT with Target加班整理:')
    ot_target_sheet = ot_wb['加班整理']
    # Dates in row 4, cols 2 to max_column-1
    target_dates = []
    for col in range(2, ot_target_sheet.max_column):
        val = ot_target_sheet.cell(4, col).value
        if isinstance(val, datetime):
            target_dates.append((col, val.strftime('%Y-%m-%d')))
        elif val and val != '加總' and val != '總計':
            target_dates.append((col, str(val).split()[0]))
            
    # Check rows 5 onwards for employees
    for r in range(5, ot_target_sheet.max_row):
        emp_name = ot_target_sheet.cell(r, 1).value
        if emp_name and emp_name != '加總' and emp_name != '總計':
            cname = get_canonical_name(emp_name)
            diffs = []
            for col_idx, date_str in target_dates:
                target_val = float(ot_target_sheet.cell(r, col_idx).value or 0)
                calc_val = calculated_ot.get(cname, {}).get(date_str, 0.0)
                if target_val != calc_val:
                    diffs.append(f'{date_str}: calc={calc_val} vs target={target_val}')
            if diffs:
                print(f'  [DIFF] {cname}: {", ".join(diffs)}')
            else:
                print(f'  [OK] {cname} matches target OT sheet!')

    # Compare 統計表 cumulative numbers
    print('\nComparing 統計表 cumulative values:')
    target_stat_sheet_name = [s for s in target_wb.sheetnames if '統計表' in s][0]
    target_stat_sheet = target_wb[target_stat_sheet_name]
    
    # Cumulative stats from leave records (Jan-May)
    cumulative_leaves = {}
    for r in leave_records:
        if r['month'] <= 5:  # up to May
            cumulative_leaves.setdefault(r['name'], {})
            cumulative_leaves[r['name']][r['type']] = cumulative_leaves[r['name']].get(r['type'], 0.0) + r['days']

    for r in range(4, target_stat_sheet.max_row+1):
        emp_name = target_stat_sheet.cell(r, 1).value
        if emp_name:
            cname = get_canonical_name(emp_name)
            # Col B = PTO, Col C = PTO-AL
            # Col P = 已休, Col Q = 剩餘
            # Col R = 病假, Col S = 生理假, Col T = paid SL, Col U = paid SL-M
            # Col V = PL, Col W = LOA, Col X = ML, Col Y = BL, Col Z = FL
            target_pto_used = float(target_stat_sheet.cell(r, 16).value or 0)  # Col P
            target_pto_remain = float(target_stat_sheet.cell(r, 17).value or 0)  # Col Q
            target_sl = float(target_stat_sheet.cell(r, 20).value or 0)  # Col T (paid SL)
            target_slm = float(target_stat_sheet.cell(r, 21).value or 0)  # Col U (paid SL-M)
            target_pl = float(target_stat_sheet.cell(r, 22).value or 0)  # Col V (PL)
            target_loa = float(target_stat_sheet.cell(r, 23).value or 0)  # Col W (LOA)

            # Calc pto used
            emp_leaves = cumulative_leaves.get(cname, {})
            calc_pto_used = emp_leaves.get('Annual Leave特別休假', 0.0) + emp_leaves.get('Asurion Leave亞勝假期', 0.0)
            
            # Calc other leaves
            calc_sl_raw = emp_leaves.get('Sick Leave病假', 0.0)
            calc_slm_raw = emp_leaves.get('Menstrual Leave病假（生理假)', 0.0)
            calc_pl = emp_leaves.get('Personal Leave事假', 0.0)
            calc_loa = sum(v for k, v in emp_leaves.items() if 'Official Leave公假' in k)
            
            # Paid SL / SL-M logic
            calc_sl = calc_sl_raw + max(calc_slm_raw - 3.0, 0.0)
            calc_slm = min(calc_slm_raw, 3.0)

            diffs = []
            if calc_pto_used != target_pto_used:
                diffs.append(f'PTO Used: calc={calc_pto_used} vs target={target_pto_used}')
            if calc_sl != target_sl:
                diffs.append(f'SL: calc={calc_sl} vs target={target_sl}')
            if calc_slm != target_slm:
                diffs.append(f'SL-M: calc={calc_slm} vs target={target_slm}')
            if calc_pl != target_pl:
                diffs.append(f'PL: calc={calc_pl} vs target={target_pl}')
            if calc_loa != target_loa:
                diffs.append(f'LOA: calc={calc_loa} vs target={target_loa}')
                
            if diffs:
                print(f'  [DIFF] {cname}: {", ".join(diffs)}')
            else:
                print(f'  [OK] {cname} matches target cumulative stats!')

if __name__ == '__main__':
    test_verification()
